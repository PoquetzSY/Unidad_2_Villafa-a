from openpyxl import Workbook,load_workbook

def excel_f():
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Ejemplo"
    sheet["B1"] = "Excel"
    workbook.save("ejemplo.xlsx")

    workbook = load_workbook("ejemplo.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row)
